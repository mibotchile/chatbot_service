#!/usr/bin/env python3
"""Repara `Numero de cuenta` corrompido (notacion cientifica de Excel) en el
PROCESADO de asignacion de Prestamype, tomando el valor LIMPIO de las fuentes
de cartera. Todo se lee y escribe como TEXTO: nunca pasa por float ni por Excel,
asi que no se vuelve a corromper.

Causa raiz (verificada 2026-06-04 en srv-ftp01 / alias ssh batch-etl-prod):
  El paso que genera Asignacion_*_PROCESADO.csv a partir de la cartera pasa por
  Excel. `Numero de cuenta` (numero largo SIN ceros a la izquierda) se convierte
  a notacion cientifica (8.98348E+12) y pierde digitos. El CCI se salva porque
  sus ceros a la izquierda lo mantienen como texto. Las cuentas reales pueden
  tener guiones y ceros a la izquierda (ej. 0011-0346-0200082747), asi que
  "limpio" = cualquier valor que NO sea notacion cientifica.

Fuente limpia (PRIORIDAD por VINTAGE — critico para no inyectar cuentas de
otro snapshot):
  1. La cartera que GENERO este PROCESADO (mismo vintage). Para el PROCESADO
     Asignacion_20052026_F es Cartera_202604_F.csv. Verificado: cubre 256/256
     de las cuentas corruptas con el valor correcto de ESE vintage.
  2. (Opcional) Otras carteras CSV mas viejas, como respaldo.
  3. (Ultimo recurso) Archivo de cartera_final.xlsx. OJO: es OTRO snapshot; sus
     numeros de cuenta pueden DIFERIR de los del PROCESADO (ej. P01250 difiere).
     Usar solo si un credito no esta en ninguna cartera del vintage.
  El mapa se arma en el ORDEN en que se pasan las fuentes (la primera gana), asi
  que pasa SIEMPRE la cartera del mismo vintage primero.

Que hace:
  - Construye {ID credito -> Numero de cuenta limpio}, xlsx primero, luego cada
    cartera CSV (sin pisar lo ya cargado). Ignora valores cientificos/vacios.
  - En el PROCESADO, reemplaza SOLO los `Numero de cuenta` en notacion cientifica
    por el valor limpio del mapa. No toca ninguna otra columna ni la logica de
    dedup/transform del PROCESADO.
  - Escribe la salida como texto (mismo delimitador ';', BOM utf-8, CRLF).

Uso (recomendado — cartera del mismo vintage, cubre 100%):
  python3 repair_numero_cuenta_prestamype.py \
      --procesado IN/Asignacion_20052026_F_PROCESADO.csv \
      --out       IN/Asignacion_20052026_F_PROCESADO_FIX.csv \
      --cartera   Cartera_202604_F.csv

  Pasa varias --cartera (la primera gana) y, solo como ultimo recurso, --xlsx
  (requiere openpyxl: `pip install openpyxl`; es otro snapshot, puede diferir).
  El script avisa (exit 1) si queda algun credito sin fuente.

Prevencion (lo importante a futuro): generar el PROCESADO con un script que lea
el xlsx con openpyxl (las celdas ya son string) y escriba `Numero de cuenta`
(y cualquier identificador numerico largo) como texto. NUNCA abrir/guardar la
cartera en Excel: ahi se rompe.
"""

from __future__ import annotations

import argparse
import csv
import re
import sys

KEY_COL = "ID credito"
FIX_COL = "Numero de cuenta"
# Un valor esta corrupto si tiene notacion cientifica ("8.98348E+12").
_SCI = re.compile(r"[eE][+-]?\d+")


def _is_scientific(value: str) -> bool:
    return bool(_SCI.search(value.strip()))


def _is_clean(value: str) -> bool:
    """Limpio = no vacio y NO cientifico (puede tener guiones / ceros a la izq.)."""
    v = value.strip()
    return bool(v) and not _is_scientific(v)


def _find(header: list[str], name: str) -> int:
    for i, h in enumerate(header):
        if h and name.lower() in str(h).strip().lower():
            return i
    sys.exit(f"ERROR: no se encontro la columna '{name}'. Header: {header}")


def add_from_xlsx(path: str, mapping: dict[str, str]) -> int:
    try:
        import openpyxl  # noqa: PLC0415
    except ModuleNotFoundError:
        sys.exit("ERROR: --xlsx requiere openpyxl. Instala con: pip install openpyxl")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = [str(h) if h is not None else "" for h in next(rows)]
    ki, ci = _find(header, KEY_COL), _find(header, FIX_COL)
    added = 0
    for row in rows:
        if ki >= len(row) or ci >= len(row):
            continue
        cred = str(row[ki]).strip() if row[ki] is not None else ""
        val = str(row[ci]).strip() if row[ci] is not None else ""
        if cred and _is_clean(val) and cred not in mapping:
            mapping[cred] = val
            added += 1
    return added


def add_from_csv(path: str, mapping: dict[str, str]) -> int:
    with open(path, newline="", encoding="utf-8-sig") as fh:
        reader = csv.reader(fh, delimiter=";")
        header = next(reader)
        ki, ci = _find(header, KEY_COL), _find(header, FIX_COL)
        added = 0
        for row in reader:
            if len(row) <= max(ki, ci):
                continue
            cred, val = row[ki].strip(), row[ci].strip()
            if cred and _is_clean(val) and cred not in mapping:
                mapping[cred] = val
                added += 1
    return added


def repair(procesado: str, out: str, mapping: dict[str, str]) -> None:
    with open(procesado, newline="", encoding="utf-8-sig") as fh:
        rows = list(csv.reader(fh, delimiter=";"))
    if not rows:
        sys.exit("ERROR: PROCESADO vacio.")
    header = rows[0]
    ki, ci = _find(header, KEY_COL), _find(header, FIX_COL)

    fixed = clean_already = unrecoverable = 0
    sin_fuente: list[str] = []
    for row in rows[1:]:
        if len(row) <= max(ki, ci):
            continue
        if not _is_scientific(row[ci]):
            clean_already += 1
            continue
        cred = row[ki].strip()
        src = mapping.get(cred)
        if src:
            row[ci] = src
            fixed += 1
        else:
            unrecoverable += 1
            sin_fuente.append(cred)

    with open(out, "w", newline="", encoding="utf-8-sig") as fh:
        csv.writer(fh, delimiter=";", quoting=csv.QUOTE_MINIMAL).writerows(rows)

    print(f"Mapa limpio combinado: {len(mapping)} creditos")
    print(f"Reparados (cientifica -> limpio): {fixed}")
    print(f"No corruptos (sin tocar):         {clean_already}")
    print(f"Irrecuperables (sin fuente):      {unrecoverable}")
    if sin_fuente:
        print("  Creditos sin valor limpio en ninguna fuente: "
              + ", ".join(sorted(set(sin_fuente))))
    print(f"Salida: {out}")
    if unrecoverable:
        sys.exit(1)  # senal para el operador: faltan fuentes


def main() -> None:
    ap = argparse.ArgumentParser(description="Repara Numero de cuenta cientifico en el PROCESADO.")
    ap.add_argument("--procesado", required=True, help="CSV PROCESADO a reparar")
    ap.add_argument("--out", required=True, help="CSV de salida reparado")
    ap.add_argument("--xlsx", default=None, help="Archivo de cartera_final.xlsx (fuente prioritaria)")
    ap.add_argument("--cartera", nargs="*", default=[], help="Cartera_*.csv de respaldo (uno o mas)")
    args = ap.parse_args()

    if not args.xlsx and not args.cartera:
        sys.exit("ERROR: pasa al menos --xlsx o --cartera como fuente limpia.")

    # Prioridad por vintage: carteras en el orden dado (la primera gana), y el
    # xlsx SOLO al final como ultimo recurso (es otro snapshot, puede diferir).
    mapping: dict[str, str] = {}
    for c in args.cartera:
        print(f"cartera {c}: +{add_from_csv(c, mapping)} creditos")
    if args.xlsx:
        print(f"xlsx {args.xlsx} (ultimo recurso): +{add_from_xlsx(args.xlsx, mapping)} creditos")
    repair(args.procesado, args.out, mapping)


if __name__ == "__main__":
    main()
